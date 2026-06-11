"""
ClearBright Digital — Lead Tracker Generator
Run this script once to produce ClearBright_Leads_Tracker.xlsx
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Raw lead data ──────────────────────────────────────────────────────────────
# Fields: Business Name, Category, Town, Phone, Email, Address, Website Status, Rating, GBP URL
LEADS = [
    ("J Sage Plumbing and Heating","Plumber","Colchester","07714 340551","","Endsleigh Ct, Colchester CO3 3QW","No Website","","https://www.google.com/maps/place/J+Sage+Plumbing+and+Heating"),
    ("GreenStar Powerflushing, Gas, Plumbing and Heating","Plumber","Colchester","07455 107724","","31A N Station Rd, Colchester CO1 1RQ","No Website","",""),
    ("Mended plumbing and heating service","Gas engineer","Colchester","07738 862960","","37 Hythe Hill, Colchester CO1 2NQ","No Website","5.0",""),
    ("Essex Plumbing & Heating","Plumber","Chelmsford","07399 446671","","Crocus Way, Springfield, Chelmsford CM1 6XJ","No Website","",""),
    ("Chelmsford Plumbers Ltd","Plumber","Chelmsford","01245 926110","","86 Linnet Dr, Chelmsford CM2 8AF","No Website","",""),
    ("All Clear Plumbing and Heating","Plumber","Chelmsford","07980 385481","","20 Further Mdw, Writtle, Chelmsford CM1 3LE","No Website","",""),
    ("Chelmsford Plumbing and Heating Services Ltd","Plumber","Chelmsford","07360 268550","","7 Goldlay Gardens, Chelmsford CM2 0EN","No Website","4.0",""),
    ("Zone 1 Plumbing","Plumber","Basildon","07821 093929","","","No Website","",""),
    ("Act Fast Plumbing","Plumber","Basildon","07842 624192","","36 Southernhay, Basildon SS14 1ET","Poor - Directory/booking platform","4.9",""),
    ("London Plumbers And Heating ltd","Central Heating Service","Basildon","07475 011972","","61 Wimbish End, Basildon SS13 3PG","No Website","",""),
    ("J Holt plumbing Ltd","Plumber","Basildon","07943 782065","","","No Website","",""),
    ("Plumbers direct 24h","Plumber","Basildon","07598 625707","","","No Website","",""),
    ("Zak's Plumbing","Plumber","Southend-on-Sea","07378 438791","","112 Brightwell Ave, Westcliff-on-Sea SS0 9EF","No Website","4.9",""),
    ("Urban Sanitation","Plumber","Southend-on-Sea","07739 719912","","42 Moseley St, Southend-on-Sea SS2 4NN","Poor - Social media page","",""),
    ("TRM Plumbing Services","Plumber","Braintree","07729 200989","","18 Dapifer Dr, Braintree CM7 3LG","Poor - Free website builder","",""),
    ("Angel Plumbing and Heating","Plumber","Braintree","01245 206143","","47 Rushmoor Dr, Braintree CM7 1TW","No Website","",""),
    ("Matt Brundle Electrical","Electrician","Colchester","07734 603202","","23 Bolsin Dr, Colchester CO4 5FD","No Website","",""),
    ("David Blundell Electrical","Electrician","Colchester","07906 868122","","10 Armidale Walk, Colchester CO2 8XS","No Website","",""),
    ("AMC Electrical","Electrical installation service","Colchester","07534 446358","","1 London Cottages, School Ln, Colchester CO6 4BP","Poor - Free website builder","",""),
    ("A B Electrical Essex Ltd","Electrician","Chelmsford","07746 833777","","Forefield Grn, Springfield, Chelmsford CM1 6YU","Poor - Social media page","",""),
    ("Mead Electrical Services Essex","Electrician","Chelmsford","07751 222603","","Tower Ave, Chelmsford CM1 2PW","Poor - Social media page","",""),
    ("MB Chelmsford Electricians","Electrician","Chelmsford","01245 331245","","Ashton Pl, Chelmer Village, Chelmsford CM2 6RF","No Website","",""),
    ("Baddow Electrical","Electrician","Chelmsford","07920 742556","","14 Longmore Ave, Great Baddow, Chelmsford CM2 7NT","No Website","",""),
    ("SS Electrical","Electrician","Basildon","07500 056818","","","Poor - Social media page","",""),
    ("AA Electrical Services","Electrician","Basildon","07800 871257","","500 Whitmore Way, Basildon SS14 2ET","No Website","5.0",""),
    ("Basildon Electrics","Electrician","Basildon","01268 525035","","14 The Fryth, Basildon SS14 3PL","No Website","5.0",""),
    ("I B Electrical London Limited","Electrician","Basildon","07894 940308","","51 Bockingham Grn, Basildon SS13 1PF","No Website","",""),
    ("JB Electrical Services","Electrician","Basildon","07455 733773","","19 Roberts Rd, Laindon, Basildon SS15 6AY","No Website","",""),
    ("Powerlec Electrical & Testing Ltd","Electrician","Southend-on-Sea","01702 952951","","39 Midhurst Ave, Westcliff-on-Sea SS0 0NP","No Website","",""),
    ("DG Electrical","Electrician","Southend-on-Sea","07799 584890","","Ladylands, Poynters Ln, Southend-on-Sea SS3 9TS","No Website","",""),
    ("Legends Hair Studio","Hairdresser","Colchester","01206 577907","","2 The Square, Iceni Way, Colchester CO2 9EB","No Website","",""),
    ("Hair Pimps","Hairdresser","Colchester","01206 710470","","15 Sir Isaac's Walk, Colchester CO1 1JJ","No Website","4.8",""),
    ("AB Beauty salon Afro-Caribbean Hair","Hairdresser","Colchester","01206 572385","","12a Headgate, Colchester CO3 3BT","No Website","4.9",""),
    ("The Chelmsford Hair Salon","Hairdresser","Chelmsford","07984 436770","","1 Tindal Square, Chelmsford CM1 1EH","Poor - Directory/booking platform","4.9",""),
    ("Hair by Rose Essex","Hairdresser","Chelmsford","07751 448051","","Scotts Walk, Chelmsford CM1 2HB","Poor - Social media page","",""),
    ("The Hair Shed","Hairdresser","Basildon","07854 046405","","Park Cottages, Bowers Ct Dr, Basildon SS13 2DT","Poor - Social media page","5.0",""),
    ("Coco Hair & Beauty","Hairdresser","Basildon","07951 838867","","30 Southernhay, Basildon SS14 1EL","No Website","",""),
    ("Hair Goddess","Hairdresser","Basildon","07456 446404","","4 Trinity Cl, Laindon, Basildon SS15 5FT","No Website","",""),
    ("Twisted hair and beauty","Hairdresser","Basildon","01268 542005","","43 Presidents Court, 50 Hoover Dr, Basildon SS15 6LF","No Website","",""),
    ("Headquarters Unisex Hair Salon","Hairdresser","Basildon","01268 543059","","29 Ballards Walk, Basildon SS15 5HL","Poor - Social media page","",""),
    ("Styles Hair Studio","Barber shop","Basildon","01268 474447","","215 Timberlog Ln, Basildon SS14 1PB","No Website","4.7",""),
    ("Reflections","Hairdresser","Basildon","01268 272737","","108 Rectory Rd, Pitsea, Basildon SS13 2AH","No Website","4.3",""),
    ("Curls and Coils Hair Salon","Hairdresser","Basildon","07940 222954","","Market Square, Basildon SS14 1BZ","No Website","",""),
    ("The Cut Inn Hairdressers","Barber shop","Basildon","01268 545310","","5 Kibcaps, Lee Chapel, Basildon SS16 5SA","No Website","",""),
    ("Head Hunters of Southend","Hairdresser","Southend-on-Sea","01702 468269","","Unit B3-B4 The Victoria Shopping Centre, Southend-on-Sea SS2 5SP","No Website","",""),
    ("Kings & Queens Unisex Salon & Barbers Southend","Hairdresser","Southend-on-Sea","01702 334400","","133 High St, Southend-on-Sea SS1 1LH","No Website","4.7",""),
    ("Reality Hair Salon","Hairdresser","Southend-on-Sea","07709 994128","","23 High St, Southend-on-Sea SS1 1JE","Poor - Social media page","4.9",""),
    ("NINETEEN HAIR & BEAUTY SALON","Hairdresser","Southend-on-Sea","01702 831511","","19 Southchurch Rd, Southend-on-Sea SS1 2NG","No Website","4.5",""),
    ("Cut U Up Hair & Beauty Salon","Hairdresser","Southend-on-Sea","01702 870601","","17 Southchurch Rd, Southend-on-Sea SS1 2NG","No Website","",""),
    ("Salon 23","Hairdresser","Southend-on-Sea","01702 817834","","23 High St, Southend-on-Sea SS1 1JE","No Website","4.5",""),
    ("Perry Browns Salon","Hairdresser","Southend-on-Sea","01702 826470","","485 Southchurch Rd, Southend-on-Sea SS1 2PH","No Website","",""),
    ("Edi's Unisex Hair Salon & Barbers","Hairdresser","Southend-on-Sea","01702 875565","","492 London Rd, Westcliff-on-Sea SS0 9LD","No Website","",""),
    ("Red Lion Cuts","Barber shop","Colchester","07918 883472","","43 High St, Colchester CO1 1DH","No Website","5.0",""),
    ("Sharp Cut","Barber shop","Colchester","Send to phone","","38 Queen St, Colchester CO1 2PL","No Website","4.8",""),
    ("Samuel - Barber","Barber shop","Colchester","07598 396883","","11 Headgate, Sir Isaac's Walk, Colchester CO3 3FH","Poor - Social media page","5.0",""),
    ("COLCH CUTZ 1","Barber shop","Colchester","07777 152234","","7 St Botolph's St, Colchester CO2 7DU","No Website","4.7",""),
    ("Top Barbers","Barber shop","Colchester","01206 642835","","74-76 Butt Rd, Colchester CO3 3DA","No Website","",""),
    ("Crown Barbers","Barber shop","Colchester","01206 765898","","8 The Centre, Iceni Way, Colchester CO2 9EB","No Website","",""),
    ("North Hill Barbers","Barber shop","Colchester","01206 547666","","30 North Hill, Colchester CO1 1EH","No Website","",""),
    ("I fade","Barber shop","Colchester","07533 224463","","33 St Botolph's St, Colchester CO2 7EA","No Website","",""),
    ("Brilliant Barber","Barber shop","Colchester","07707 929172","","131A High St, Colchester CO1 1SP","No Website","",""),
    ("Mews Barbers","Barber shop","Colchester","01206 572664","","2A Church St, Colchester CO1 1NF","Poor - Social media page","",""),
    ("Hair Razor Barbers","Barber shop","Colchester","07949 993766","","20 Eld Ln, Colchester CO1 1LS","No Website","",""),
    ("The Man Cave","Barber shop","Chelmsford","01245 611523","","1 High St, Chelmsford CM1 1BE","Poor - Social media page","4.8",""),
    ("Essex Barbers Lounge","Barber shop","Chelmsford","01245 350777","","31 Broomfield Rd, Chelmsford CM1 1SY","No Website","",""),
    ("Diverse Barbershop - Chelmsford","Barber shop","Chelmsford","07376 421024","","12a Parkway, Chelmsford CM2 0NF","Poor - Directory/booking platform","5.0",""),
    ("Moulsham Barber's Lounge","Barber shop","Chelmsford","01245 261700","","161A Moulsham St, Chelmsford CM2 0LD","Poor - Directory/booking platform","4.8",""),
    ("K1 BARBERS","Barber shop","Chelmsford","07982 395945","","7 Hylands Parade, Wood St, Chelmsford CM2 8BW","No Website","",""),
    ("H4Z FADEZ","Barber shop","Chelmsford","07961 991847","","66 Forest Dr, Chelmsford CM1 2TS","No Website","5.0",""),
    ("The Men's Salon","Barber shop","Chelmsford","01245 283840","","Kiosk 1, Meadows Shopping Centre, Chelmsford CM2 6FD","No Website","4.8",""),
    ("Old town barbers","Barber shop","Basildon","01268 525951","","58 E Walk, Basildon SS14 1HE","No Website","4.2",""),
    ("Hot Shots Barbers Basildon","Barber shop","Basildon","01268 287111","","17 Eastgate, Basildon SS14 1JJ","No Website","4.4",""),
    ("Tuff Luck Barbers","Barber shop","Basildon","01268 282870","","Unit 1 Paycocke Rd, Basildon SS14 3EH","Poor - Directory/booking platform","",""),
    ("Magic Scissors","Barber shop","Basildon","01268 906363","","83 Southernhay, Basildon SS14 1EU","Poor - Social media page","3.9",""),
    ("Crown turkish barber in kingswood","Barber shop","Basildon","07466 031457","","142 Clay Hill Rd, Basildon SS16 5DF","No Website","",""),
    ("Istanbul","Barber shop","Basildon","01268 952851","","410 Whitmore Way, Basildon SS14 2HB","Poor - Social media page","3.6",""),
    ("Stocks Barbershop","Barber shop","Basildon","07951 171401","","Burnt Mills Rd, Basildon SS13 1DY","No Website","",""),
    ("Moe's barbers","Barber shop","Basildon","01268 928798","","South Walk, 10 S Walk, Basildon SS14 1BZ","No Website","",""),
    ("Pitsea Turkish Barber","Barber shop","Basildon","07411 165922","","Pitsea, Basildon SS13 3BG","Poor - Social media page","",""),
    ("Kings Turkish Barbers pitsea","Barber shop","Basildon","07440 125498","","63 High Rd, Pitsea, Basildon SS13 3BB","No Website","",""),
    ("Laindon Turkish barber","Barber shop","Basildon","07424 847919","","High Road Laindon, Basildon SS15 6NR","No Website","",""),
    ("Felmores Barber","Barber shop","Basildon","01268 217206","","5 Felmores End, Basildon SS13 1PN","No Website","",""),
    ("String's HQ","Barber shop","Basildon","01268 534880","","Unit 36b, Eastgate shopping centre, Basildon SS14 1AE","No Website","",""),
    ("Jimmys Barbers","Barber shop","Southend-on-Sea","07703 755185","","141 Southchurch Rd, Southend-on-Sea SS1 2NW","Poor - Social media page","4.9",""),
    ("Mo's Barber","Barber shop","Southend-on-Sea","07411 311806","","633B London Rd, Westcliff-on-Sea SS0 9PE","No Website","4.8",""),
    ("Southend barber","Barber shop","Southend-on-Sea","07476 753477","","151 Fairfax Dr, Westcliff-on-Sea SS0 9BQ","No Website","",""),
    ("Zack Barbers","Barber shop","Southend-on-Sea","07429 363631","","99 Southchurch Rd, Southend-on-Sea SS1 2NL","No Website","4.9",""),
    ("Arastocuts","Barber shop","Southend-on-Sea","01702 831820","","102 High St, Southend-on-Sea SS1 1JN","No Website","4.6",""),
    ("Mo's barber Southend","Barber shop","Southend-on-Sea","07480 810706","","12 Alexandra St, Southend-on-Sea SS1 1BU","No Website","4.7",""),
    ("Cutz & Bladez","Barber shop","Southend-on-Sea","07454 545895","","104 Southchurch Rd, Southend-on-Sea SS1 2LX","No Website","4.7",""),
    ("Men Only","Barber shop","Southend-on-Sea","07966 587295","","30 Woodcutters Ave, Leigh-on-Sea SS9 4PL","No Website","",""),
    ("Rag N Bone Barber Co","Barber shop","Southend-on-Sea","07848 685415","","50 Hamlet Ct Rd, Westcliff-on-Sea SS0 7LX","No Website","4.9",""),
    ("Roots & Grooves Cafe","Cafe","Colchester","Send to phone","","1 St Nicholas Passage, Colchester CO1 1TB","No Website","4.7",""),
    ("Ella's Vietnamese Kitchen","Cafe","Colchester","01206 626508","","20 Sir Isaac's Walk, Colchester CO1 1JJ","No Website","4.9",""),
    ("Sir Isaac's Artisan Coffee","Coffee shop","Colchester","01206 803105","","43 Sir Isaac's Walk, Colchester CO1 1JJ","No Website","4.4",""),
    ("Cafe Local Colchester","Convenience Store","Colchester","01206 573440","","Railway Station, N Station Rd, Colchester CO1 1XD","No Website","",""),
    ("Buzzcoffee","Coffee shop","Colchester","07453 289856","","59B N Station Rd, Colchester CO1 1RQ","No Website","",""),
    ("Chicago's Coffee & Sandwich Bar","Restaurant","Colchester","01206 417437","","8 Eld Ln, Colchester CO1 1LS","No Website","",""),
    ("Jenkins Cafe","Cafe","Colchester","01206 521448","","42 St John's St, Colchester CO2 7AD","Poor - Social media page","4.6",""),
    ("The Brunch Café","Cafe","Colchester","01206 586705","","Colchester CO3 0LA","Poor - Social media page","",""),
    ("Central Park Cafe Chelmsford","Cafe","Chelmsford","01245 931590","","Central Park, Chelmsford CM1 1LQ","Poor - Social media page","4.6",""),
    ("The Little Cafe","Cafe","Chelmsford","01245 359797","","201a Moulsham St, Chelmsford CM2 0LG","No Website","4.7",""),
    ("Niffer's Café chelmsford","Restaurant","Chelmsford","07958 452366","","Inside Selco Builders Merchant, Chelmsford CM1 3AR","No Website","",""),
    ("The Cafe by Benugo at John Lewis","Cafe","Chelmsford","01245 458400","","Bond St, Chelmsford CM1 1GD","No Website","",""),
    ("House Cleaning Services Colchester","House cleaning service","Colchester","01202 464586","","115 N Station Rd, Colchester CO1 1SB","No Website","",""),
    ("Shine Sharpers Limited","House cleaning service","Chelmsford","07309 114380","","Haig Ct, Chelmsford CM2 0BJ","No Website","5.0",""),
    ("Samas cleaning Ltd","Cleaners","Basildon","07400 063949","","Academy Dr, Basildon SS15 6GQ","No Website","",""),
    ("Affordable Cleaning Company","Cleaning service","Basildon","07891 904303","","30 Stokefelde, Pitsea, Basildon SS13 1NH","No Website","",""),
    ("VV touch of perfection housekeeping ltd","House cleaning service","Basildon","07949 163202","","Darwin Ct, Basildon SS14 3SY","No Website","",""),
    ("DC PRO Solutions Ltd","Carpet cleaning service","Basildon","Send to phone","","","No Website","",""),
    ("Pixie Perfect Cleaning & Co","House cleaning service","Basildon","07984 009712","","","No Website","",""),
    ("Sol's Decorating - Painting and decorating","Painter and Decorator","Colchester","07795 213444","","28 Peache Rd, Colchester CO1 2FS","No Website","5.0",""),
    ("Deluxe Painters And Decorators","Painter and Decorator","Colchester","07421 877992","","Grantham Ct, Colchester CO1 2RU","Poor - Social media page","",""),
    ("B P Decorating Services","Painter and Decorator","Colchester","01206 530703","","4 Hamilton Rd, Colchester CO3 3DZ","No Website","5.0",""),
    ("Wise Painting and Decorating","Painter and Decorator","Colchester","07300 346606","","","Poor - Social media page","5.0",""),
    ("SP COLLYER DECORATING","Painter and Decorator","Colchester","07961 800535","","29 Steed Cres, Colchester CO2 7SJ","No Website","",""),
    ("Painter & Decorator","Painter and Decorator","Colchester","07834 640570","","46 Broadlands Way, Colchester CO4 0AN","No Website","5.0",""),
    ("Painter and Decorator Colchester","Painter and Decorator","Colchester","020 3906 1182","","5 Whitehall Cl, Colchester CO2 8AJ","No Website","",""),
    ("TJP Decorating","Interior Decorator","Chelmsford","07557 911044","","Louis Faiers Cres, Boreham, Chelmsford CM1 6EY","No Website","",""),
    ("D'Decor Ltd","Painter and Decorator","Chelmsford","Send to phone","","","Poor - Social media page","",""),
    ("Hayes Paul","Painter and Decorator","Chelmsford","01245 463081","","3 Wavell Cl, Springfield, Chelmsford CM1 6FQ","No Website","",""),
    ("Rgl Painter & Decorator","Painter and Decorator","Basildon","07976 720438","","5 Clarendon Rd, Basildon SS13 2BL","No Website","5.0",""),
    ("Pro touch painting and decorating services","Painter and Decorator","Basildon","07493 221534","","30 Lippits Hill, Langdon Hills, Basildon SS16 6LN","No Website","5.0",""),
    ("Painters in Basildon","Painter and Decorator","Basildon","07400 163236","","St Andrews Ln, Laindon, Basildon SS15 5WH","No Website","",""),
    ("M. A. Smith Building and Decorating Services","Painter and Decorator","Basildon","01268 417418","","12 Beecham Ct, Basildon SS15 5RB","No Website","5.0",""),
    ("CP WILSON PAINTING & DECORATING","Painter and Decorator","Basildon","07966 172080","","26 Wellstye Grn, Basildon SS14 2SR","No Website","",""),
    ("Painter & Decorator Basildon - RS Homes","Painter and Decorator","Basildon","07771 864011","","46 Little Oxcroft, Laindon, Basildon SS15 6PD","No Website","",""),
    ("vibrant visions painting and decorating","Interior Decorator","Basildon","Send to phone","","135 Swanstead, Basildon SS16 4PA","No Website","5.0",""),
    ("HollinsDecoratingLtd","Painter and Decorator","Basildon","07718 935832","","22 Heathleigh Dr, Langdon Hills, Basildon SS16 6AR","No Website","5.0",""),
]

# ── Style helpers ──────────────────────────────────────────────────────────────
CHARCOAL, GOLD, WHITE = "1A1A1A", "C89B3C", "FFFFFF"
LIGHT_GREY, MID_GREY  = "F5F5F5", "E0E0E0"
GREEN_BG,  GREEN_FG   = "D6EFD8", "1A5C1A"
AMBER_BG,  AMBER_FG   = "FFF3CD", "7D5A00"

def hdr_font(**kw):  return Font(name="Arial", bold=True,  size=9,  color=WHITE,   **kw)
def dat_font(**kw):  return Font(name="Arial", bold=False, size=9,  color=CHARCOAL, **kw)
def bold_font(**kw): return Font(name="Arial", bold=True,  size=9,  color=CHARCOAL, **kw)
def solid(hex_):     return PatternFill("solid", fgColor=hex_)
def border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ── Column definitions ─────────────────────────────────────────────────────────
COLS = [
    ("#",               5,  "center"),
    ("Business Name",   32, "left"),
    ("Category",        20, "left"),
    ("Town",            14, "left"),
    ("Phone",           16, "left"),
    ("Email",           24, "left"),
    ("Address",         30, "left"),
    ("Website Status",  24, "left"),
    ("Rating",          8,  "center"),
    ("Outreach Status", 20, "center"),
    ("Email Sent",      14, "center"),
    ("Follow-up 1",     14, "center"),
    ("Follow-up 2",     14, "center"),
    ("Notes",           32, "left"),
]

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Leads"
ws.freeze_panes = "A3"

# Row 1 — title banner
last_col = get_column_letter(len(COLS))
ws.merge_cells(f"A1:{last_col}1")
t = ws["A1"]
t.value = "ClearBright Digital — Lead Tracker"
t.font  = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill  = solid(CHARCOAL)
t.alignment = align("center")
ws.row_dimensions[1].height = 28

# Row 2 — column headers
for ci, (hdr, width, ha) in enumerate(COLS, 1):
    c = ws.cell(row=2, column=ci, value=hdr)
    c.font      = hdr_font()
    c.fill      = solid(GOLD)
    c.alignment = align("center")
    c.border    = border()
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[2].height = 20

# Data rows
for ri, lead in enumerate(LEADS, 3):
    biz, cat, town, phone, email, address, web, rating, _ = lead
    row_fill = solid(WHITE) if ri % 2 else solid(LIGHT_GREY)

    values = [
        ri - 2, biz, cat, town, phone, email, address,
        web,
        rating if rating not in ("", "0.0") else "",
        "Not Started", "", "", "", "",
    ]

    for ci, val in enumerate(values, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font   = dat_font()
        c.fill   = row_fill
        c.border = border()
        _, _, ha = COLS[ci - 1]
        c.alignment = align(ha, wrap=(ci in [2, 7, 14]))

    # Website status colour
    wc = ws.cell(row=ri, column=8)
    if web == "No Website":
        wc.fill = solid(GREEN_BG)
        wc.font = Font(name="Arial", size=9, color=GREEN_FG)
    elif "Poor" in web:
        wc.fill = solid(AMBER_BG)
        wc.font = Font(name="Arial", size=9, color=AMBER_FG)

    ws.row_dimensions[ri].height = 18

last_row = 2 + len(LEADS)

# Dropdown for Outreach Status
dv = DataValidation(
    type="list",
    formula1='"Not Started,Email Sent,Follow-up 1 Sent,Follow-up 2 Sent,Responded,Meeting Booked,Won,Not Interested"',
    showDropDown=False
)
ws.add_data_validation(dv)
dv.add(f"J3:J{last_row}")

# Date format for date columns
for row in range(3, last_row + 1):
    for col in [11, 12, 13]:
        ws.cell(row=row, column=col).number_format = "DD/MM/YYYY"

# Auto-filter
ws.auto_filter.ref = f"A2:{last_col}{last_row}"

# ── Summary sheet ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "ClearBright Digital — Lead Summary"
t2.font  = Font(name="Arial", bold=True, size=14, color=WHITE)
t2.fill  = solid(CHARCOAL)
t2.alignment = align("center")
ws2.row_dimensions[1].height = 28

def s2h(cell, text):
    cell.value = text
    cell.font  = hdr_font()
    cell.fill  = solid(GOLD)
    cell.alignment = align("center")
    cell.border = border()

def s2v(cell, val):
    cell.value = val
    cell.font  = dat_font()
    cell.alignment = align("center")
    cell.border = border()

# By Town
ws2.cell(row=3, column=1, value="Leads by Town").font = bold_font(size=11)
s2h(ws2["A4"], "Town"); s2h(ws2["B4"], "Count")
towns = sorted(set(r[2] for r in LEADS))
for i, town in enumerate(towns, 5):
    ws2.cell(row=i, column=1, value=town).font = dat_font()
    ws2.cell(row=i, column=1).border = border()
    ws2.cell(row=i, column=1).alignment = align("left")
    s2v(ws2.cell(row=i, column=2), f"=COUNTIF(Leads!D:D,A{i})")
total_r = 5 + len(towns)
ws2.cell(row=total_r, column=1, value="TOTAL").font = bold_font()
ws2.cell(row=total_r, column=1).border = border()
ws2.cell(row=total_r, column=1).fill = solid(MID_GREY)
ws2.cell(row=total_r, column=1).alignment = align("left")
s2v(ws2.cell(row=total_r, column=2), f"=SUM(B5:B{total_r-1})")
ws2.cell(row=total_r, column=2).font = bold_font()
ws2.cell(row=total_r, column=2).fill = solid(MID_GREY)

# By Category
ws2.cell(row=3, column=4, value="Leads by Category").font = bold_font(size=11)
s2h(ws2["D4"], "Category"); s2h(ws2["E4"], "Count"); s2h(ws2["F4"], "No Website")
cats = sorted(set(r[1] for r in LEADS))
for i, cat in enumerate(cats, 5):
    ws2.cell(row=i, column=4, value=cat).font = dat_font()
    ws2.cell(row=i, column=4).border = border()
    ws2.cell(row=i, column=4).alignment = align("left")
    s2v(ws2.cell(row=i, column=5), f"=COUNTIF(Leads!C:C,D{i})")
    s2v(ws2.cell(row=i, column=6), f'=COUNTIFS(Leads!C:C,D{i},Leads!H:H,"No Website")')
cat_total = 5 + len(cats)
for col, val in [(4, "TOTAL"), (5, f"=SUM(E5:E{cat_total-1})"), (6, f"=SUM(F5:F{cat_total-1})")]:
    c = ws2.cell(row=cat_total, column=col, value=val)
    c.font   = bold_font()
    c.border = border()
    c.fill   = solid(MID_GREY)
    c.alignment = align("left" if col == 4 else "center")

# Col widths
for col, w in [("A",28),("B",10),("C",4),("D",28),("E",10),("F",14)]:
    ws2.column_dimensions[col].width = w
for r in range(1, cat_total + 2):
    ws2.row_dimensions[r].height = 20

# ── Save ───────────────────────────────────────────────────────────────────────
import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ClearBright_Leads_Tracker.xlsx")
wb.save(out)
print(f"✓ Saved: {out}")
print(f"  {len(LEADS)} leads across {len(towns)} towns and {len(cats)} categories")
